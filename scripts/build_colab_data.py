"""Build compact 30-year workbooks for GitHub and Colab."""

from pathlib import Path
import sys

from openpyxl import Workbook, load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

from configs.ultimo_2025 import ASSET_FILE, ASSET_SHEETS, BEI_FILE, SWAP_FILE


OUTPUT_DIR = PROJECT_ROOT / "data" / "colab"
ASSET_OUTPUT = OUTPUT_DIR / "ultimo_2025_assets_30y.xlsx"
SWAP_OUTPUT = OUTPUT_DIR / "ultimo_2025_swap_30y.xlsx"
BEI_OUTPUT = OUTPUT_DIR / "ultimo_2025_bei_30y.xlsx"
DIRECT_LENDING_FILE = PROJECT_ROOT / "data" / "60_jaars_Flags & Frictions_Dec25.xlsx"
DIRECT_LENDING_SHEET = "8. Rendement Direct Lending Eur"

YEARS = 30
SCENARIOS = 2000
ASSET_SKIPROWS = 18
CURVE_SKIPROWS = 3

# The loader reads data from column C onward. Copy A:B as context and C:AF as the
# first 30 usable year/tenor columns.
MAX_COLUMN = 2 + YEARS


OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


print(f"Writing {ASSET_OUTPUT}")
source_wb = load_workbook(ASSET_FILE, read_only=True, data_only=True)
direct_lending_wb = load_workbook(DIRECT_LENDING_FILE, read_only=True, data_only=True)
target_wb = Workbook(write_only=True)

asset_sheet_names = ["Table of contents", *dict.fromkeys(ASSET_SHEETS.values())]
for sheet_name in asset_sheet_names:
    source_ws = source_wb[sheet_name]
    target_ws = target_wb.create_sheet(sheet_name)
    max_row = 1 + ASSET_SKIPROWS + SCENARIOS if sheet_name == "Table of contents" else ASSET_SKIPROWS + SCENARIOS
    for row in source_ws.iter_rows(
        min_row=1,
        max_row=max_row,
        min_col=1,
        max_col=MAX_COLUMN,
        values_only=True,
    ):
        target_ws.append(row)

    if sheet_name == ASSET_SHEETS["Cash"]:
        source_ws = direct_lending_wb[DIRECT_LENDING_SHEET]
        target_ws = target_wb.create_sheet(DIRECT_LENDING_SHEET)
        for row in source_ws.iter_rows(
            min_row=1,
            max_row=ASSET_SKIPROWS + SCENARIOS,
            min_col=1,
            max_col=MAX_COLUMN,
            values_only=True,
        ):
            target_ws.append(row)

target_wb.save(ASSET_OUTPUT)
source_wb.close()
direct_lending_wb.close()
print(f"Saved {ASSET_OUTPUT} ({ASSET_OUTPUT.stat().st_size / 1024 / 1024:.1f} MB)")


for source_path, output_path, prefix in (
    (SWAP_FILE, SWAP_OUTPUT, "NOM"),
    (BEI_FILE, BEI_OUTPUT, "BEI"),
):
    print(f"Writing {output_path}")
    source_wb = load_workbook(source_path, read_only=True, data_only=True)
    target_wb = Workbook(write_only=True)

    for year in range(YEARS):
        sheet_name = f"{prefix} {year}"
        source_ws = source_wb[sheet_name]
        target_ws = target_wb.create_sheet(sheet_name)
        for row in source_ws.iter_rows(
            min_row=1,
            max_row=CURVE_SKIPROWS + SCENARIOS,
            min_col=1,
            max_col=MAX_COLUMN,
            values_only=True,
        ):
            target_ws.append(row)

    target_wb.save(output_path)
    source_wb.close()
    print(f"Saved {output_path} ({output_path.stat().st_size / 1024 / 1024:.1f} MB)")
