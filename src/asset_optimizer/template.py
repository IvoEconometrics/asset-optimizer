"""Create and read portfolio input workbooks."""

from pathlib import Path
from shutil import copyfile

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from asset_optimizer.loader import ScenarioSet


RATE_COMPONENT = "Rente"
GROUP_KEYS = tuple(f"Group_{idx}" for idx in range(1, 6))
METRIC_LABELS = {
    "Weight_Sum": "Som gewichten",
    "Arith_Return": "Rekenkundig rendement",
    "Geo_Return": "Geometrisch rendement",
    "Volatility": "Volatiliteit",
    "Compound_Vol": "Volatiliteit totaal rendement",
    "Sharpe": "Sharpe ratio",
    "Pct_Pos_Return": "Percentage positief rendement",
    "Pct_Neg_Excess_Terminal": "Percentage negatief totaal overrendement",
    "Pct_Neg_Excess_Overall": "Percentage negatief jaarlijks overrendement",
    "Mean_Neg_Excess_Overall": "Gemiddeld negatief jaarlijks overrendement",
}
PERCENT_METRICS = {
    "Weight_Sum",
    "Arith_Return",
    "Geo_Return",
    "Volatility",
    "Compound_Vol",
    "Pct_Pos_Return",
    "Pct_Neg_Excess_Terminal",
    "Pct_Neg_Excess_Overall",
    "Mean_Neg_Excess_Overall",
}
PERCENT_POINT_METRICS = {
    "Pct_Pos_Return",
    "Pct_Neg_Excess_Terminal",
    "Pct_Neg_Excess_Overall",
}


def create_input_template(
    path: str | Path,
    scenario_set: ScenarioSet,
    *,
    tenor_count: int = 120,
) -> Path:
    """Create the workbook that a portfolio manager fills in."""

    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    input_fill = PatternFill("solid", fgColor="FFF2CC")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    bold = Font(bold=True)
    components = list(scenario_set.asset_names)
    if scenario_set.yields is not None:
        components.append(RATE_COMPONENT)

    wb = Workbook()
    wb.remove(wb.active)

    # Uitleg
    ws = wb.create_sheet("uitleg")
    explanation_rows = [
        ("Dit is het input Excel bestand voor de Scenario Optimizer.", False),
        ("", False),
        ("Het bestand bestaat uit tabbladen voor verschillende doeleinden.", False),
        ("", False),
        ("Portfolios:", True),
        (
            'Tabblad "portfolios" wordt gebruikt om vooraf geconstrueerde portfolios door te rekenen. '
            "De kolommen kunnen aangepast worden naar voorkeur van de gebruiker. De gewichten in iedere "
            "portefeuillekolom moeten optellen tot 1 en weerspiegelen de verdeling van de assets.",
            False,
        ),
        ("", False),
        ("Benchmark:", True),
        (
            'Tabblad "benchmark" bepaalt of rendementen absoluut of ten opzichte van een benchmark worden '
            "bekeken. Als bijvoorbeeld Prijsinflatie NL op 1 staat, worden de statistieken en de frontier "
            "berekend op basis van overrendement ten opzichte van inflatie. Zonder benchmark worden de "
            "portfolios standaard absoluut doorgerekend.",
            False,
        ),
        ("", False),
        ("Optimization:", True),
        (
            'Tabblad "optimization" bevat de bandbreedtes voor assets en groepen. In de Min- en Max-kolommen '
            "staan de toegestane gewichten. De groepenkolommen kunnen gebruikt worden voor restricties op "
            "mandjes van assets, bijvoorbeeld een zakelijke waarden restrictie.",
            False,
        ),
        ("", False),
        ("Output:", True),
        (
            'In de outputbestanden staat een extra tabblad "output". Daarin staan de doorgerekende '
            "portefeuillegewichten met direct daaronder de statistieken. De frontier-punten staan in "
            "dezelfde tabel achter de portefeuilles. De frontier-plot wordt in de notebook gemaakt.",
            False,
        ),
    ]
    for row_idx, (text, is_heading) in enumerate(explanation_rows, start=1):
        cell = ws.cell(row=row_idx, column=1, value=text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if is_heading:
            cell.font = Font(bold=True, size=12)
        elif row_idx == 1:
            cell.font = Font(size=12)
    ws.column_dimensions["A"].width = 95
    for row_idx in range(1, len(explanation_rows) + 1):
        ws.row_dimensions[row_idx].height = 32

    # Settings
    ws = wb.create_sheet("settings")
    rows = [
        ("Setting", "Value", "Explanation"),
        ("horizon_years", scenario_set.horizon_years, "Minimum horizon is 5 years."),
        ("step_size", 0.025, "Portfolio grid step. Example: 0.025 means 2.5%."),
        ("frontier_points", 250, "Internal frontier sweep count; normally leave unchanged."),
    ]
    for row_idx, row in enumerate(rows, start=1):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    for col_idx in range(1, 4):
        ws.cell(row=1, column=col_idx).font = bold
        ws.cell(row=1, column=col_idx).fill = header_fill
    for row in ws.iter_rows(min_row=2, max_row=4, min_col=2, max_col=2):
        for cell in row:
            cell.fill = input_fill
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 72

    # Cashflow is only needed when the config loaded yield curves, because it constructs Rente.
    if scenario_set.yields is not None:
        ws = wb.create_sheet("cashflow")
        ws.cell(row=1, column=1, value="Tenor")
        ws.cell(row=1, column=2, value="Weight")
        tenors = scenario_set.yield_tenors or list(range(1, tenor_count + 1))
        for row_idx, tenor in enumerate(tenors, start=2):
            ws.cell(row=row_idx, column=1, value=tenor)
            ws.cell(row=row_idx, column=2, value=0.0)
        for col_idx in range(1, 3):
            ws.cell(row=1, column=col_idx).font = bold
            ws.cell(row=1, column=col_idx).fill = header_fill
        for row in ws.iter_rows(min_row=2, max_row=len(tenors) + 1, min_col=2, max_col=2):
            for cell in row:
                cell.fill = input_fill
        ws.freeze_panes = "A2"
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 18

    # Benchmark
    ws = wb.create_sheet("benchmark")
    ws.cell(row=1, column=1, value="Component")
    ws.cell(row=1, column=2, value="Weight")
    for row_idx, component in enumerate(components, start=2):
        ws.cell(row=row_idx, column=1, value=component)
        ws.cell(row=row_idx, column=2, value=0.0)
    for col_idx in range(1, 3):
        ws.cell(row=1, column=col_idx).font = bold
        ws.cell(row=1, column=col_idx).fill = header_fill
    for row in ws.iter_rows(min_row=2, max_row=len(components) + 1, min_col=2, max_col=2):
        for cell in row:
            cell.fill = input_fill
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 18

    # Optimization bounds and groups
    ws = wb.create_sheet("optimization")
    optimization_headers = ("Component", "Min", "Max", *GROUP_KEYS)
    for col_idx, value in enumerate(optimization_headers, start=1):
        ws.cell(row=1, column=col_idx, value=value)
        ws.cell(row=1, column=col_idx).font = bold
        ws.cell(row=1, column=col_idx).fill = header_fill
    for row_idx, component in enumerate(components, start=2):
        ws.cell(row=row_idx, column=1, value=component)
    for row in ws.iter_rows(
        min_row=2,
        max_row=len(components) + 1,
        min_col=2,
        max_col=len(optimization_headers),
    ):
        for cell in row:
            cell.fill = input_fill

    group_start = len(components) + 4
    group_headers = ("Group_Key", "Group_Name", "Min", "Max")
    for col_idx, value in enumerate(group_headers, start=1):
        ws.cell(row=group_start, column=col_idx, value=value)
        ws.cell(row=group_start, column=col_idx).font = bold
        ws.cell(row=group_start, column=col_idx).fill = header_fill
    for offset, group_key in enumerate(GROUP_KEYS, start=1):
        ws.cell(row=group_start + offset, column=1, value=group_key)
    for row in ws.iter_rows(
        min_row=group_start + 1,
        max_row=group_start + len(GROUP_KEYS),
        min_col=2,
        max_col=4,
    ):
        for cell in row:
            cell.fill = input_fill
    ws.freeze_panes = "A2"
    for col_idx, width in enumerate([34, 14, 14, 14, 14, 14, 14, 14], start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Fixed portfolios to evaluate
    ws = wb.create_sheet("portfolios")
    portfolio_headers = ("Component", "Portfolio_1", "Portfolio_2", "Portfolio_3")
    for col_idx, value in enumerate(portfolio_headers, start=1):
        ws.cell(row=1, column=col_idx, value=value)
        ws.cell(row=1, column=col_idx).font = bold
        ws.cell(row=1, column=col_idx).fill = header_fill
    for row_idx, component in enumerate(components, start=2):
        ws.cell(row=row_idx, column=1, value=component)
    for row in ws.iter_rows(
        min_row=2,
        max_row=len(components) + 1,
        min_col=2,
        max_col=len(portfolio_headers),
    ):
        for cell in row:
            cell.fill = input_fill
    ws.freeze_panes = "A2"
    for col_idx, width in enumerate([34, 16, 16, 16], start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(output_path)
    return output_path


def load_input_template(path: str | Path, scenario_set: ScenarioSet) -> dict[str, pd.Series | pd.DataFrame]:
    """Read a filled portfolio input workbook into pandas objects."""

    path = Path(path)
    components = list(scenario_set.asset_names)
    if scenario_set.yields is not None:
        components.append(RATE_COMPONENT)
    allowed_components = set(components)

    workbook = pd.ExcelFile(path)

    settings_df = pd.read_excel(workbook, sheet_name="settings")
    settings = settings_df.set_index("Setting")["Value"]

    if "cashflow" in workbook.sheet_names:
        cashflow_df = pd.read_excel(workbook, sheet_name="cashflow")
        cashflow = pd.to_numeric(cashflow_df["Weight"], errors="coerce").fillna(0.0)
        cashflow.index = pd.to_numeric(cashflow_df["Tenor"], errors="raise").astype(int)
        cashflow.name = "Weight"
    else:
        cashflow = pd.Series(dtype=float, name="Weight")

    benchmark_df = pd.read_excel(workbook, sheet_name="benchmark")
    benchmark = pd.to_numeric(benchmark_df["Weight"], errors="coerce").fillna(0.0)
    benchmark.index = benchmark_df["Component"].astype(str)
    benchmark.name = "Weight"
    unknown = sorted(set(benchmark.index) - allowed_components)
    if unknown:
        raise ValueError(f"Unknown benchmark components: {unknown}")

    optimization_df = pd.read_excel(workbook, sheet_name="optimization")
    group_marker = optimization_df.index[optimization_df["Component"].eq("Group_Key")]
    if len(group_marker) != 1:
        raise ValueError("Optimization sheet must contain one Group_Key section.")

    group_marker = int(group_marker[0])
    component_rows = optimization_df.iloc[:group_marker].dropna(subset=["Component"]).copy()
    component_rows["Component"] = component_rows["Component"].astype(str)
    unknown = sorted(set(component_rows["Component"]) - allowed_components)
    if unknown:
        raise ValueError(f"Unknown optimization components: {unknown}")

    component_rows = component_rows.set_index("Component")
    asset_bounds = component_rows[["Min", "Max"]].apply(pd.to_numeric, errors="coerce")
    group_memberships = component_rows[list(GROUP_KEYS)]

    group_rows = optimization_df.iloc[group_marker + 1 :].dropna(subset=["Component"]).copy()
    group_bounds = pd.DataFrame(index=group_rows["Component"].astype(str))
    group_bounds.index.name = "Group_Key"
    group_bounds["Group_Name"] = group_rows["Min"].to_numpy()
    group_bounds["Min"] = pd.to_numeric(group_rows["Max"], errors="coerce").to_numpy()
    group_bounds["Max"] = pd.to_numeric(group_rows["Group_1"], errors="coerce").to_numpy()

    portfolios = pd.read_excel(workbook, sheet_name="portfolios").set_index("Component")
    portfolios.index = portfolios.index.astype(str)
    unknown = sorted(set(portfolios.index) - allowed_components)
    if unknown:
        raise ValueError(f"Unknown portfolio components: {unknown}")
    portfolios = portfolios.apply(pd.to_numeric, errors="coerce").fillna(0.0)

    return {
        "settings": settings,
        "cashflow": cashflow,
        "benchmark": benchmark,
        "portfolios": portfolios,
        "asset_bounds": asset_bounds,
        "group_memberships": group_memberships,
        "group_bounds": group_bounds,
    }


def write_output_workbook(
    input_path: str | Path,
    output_path: str | Path,
    portfolios: pd.DataFrame,
    portfolio_stats: pd.DataFrame,
    frontier: pd.DataFrame | None = None,
) -> Path:
    """Copy an input workbook and add one portfolio report sheet named ``output``."""

    input_path = Path(input_path)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    copyfile(input_path, output_path)

    wb = load_workbook(output_path)
    if "output" in wb.sheetnames:
        del wb["output"]
    ws = wb.create_sheet("output")

    title_fill = PatternFill("solid", fgColor="1F4E78")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    metric_fill = PatternFill("solid", fgColor="E2F0D9")
    white_bold = Font(bold=True, color="FFFFFF")
    bold = Font(bold=True)

    weights = portfolios.copy().fillna(0.0)
    weights.index.name = "Component"
    stats = portfolio_stats.copy()
    if "Portfolio" in stats.columns:
        stats = stats.set_index("Portfolio")

    if frontier is not None and not frontier.empty:
        frontier_names = [f"Frontier {idx}" for idx in range(1, len(frontier) + 1)]
        frontier_weights = frontier.reindex(columns=weights.index).T.fillna(0.0)
        frontier_weights.columns = frontier_names
        weights = pd.concat([weights, frontier_weights], axis=1)

        frontier_stats = pd.DataFrame(index=frontier_names)
        frontier_stats["Weight_Sum"] = frontier_weights.sum(axis=0)
        for metric in METRIC_LABELS:
            if metric in frontier.columns:
                frontier_stats[metric] = frontier[metric].to_numpy(dtype=float)
        stats = pd.concat([stats, frontier_stats], axis=0)

    stats = stats.reindex(weights.columns)
    metric_rows = [
        metric
        for metric in (
            "Weight_Sum",
            "Arith_Return",
            "Geo_Return",
            "Volatility",
            "Compound_Vol",
            "Sharpe",
            "Pct_Pos_Return",
            "Pct_Neg_Excess_Terminal",
            "Pct_Neg_Excess_Overall",
            "Mean_Neg_Excess_Overall",
        )
        if metric in stats.columns
    ]

    ws.cell(row=1, column=1, value=f"Portfolio output - {input_path.name}")
    ws.cell(row=1, column=1).font = white_bold
    ws.cell(row=1, column=1).fill = title_fill

    header_row = 3
    ws.cell(row=header_row, column=1, value="Component / statistiek")
    ws.cell(row=header_row, column=1).font = bold
    ws.cell(row=header_row, column=1).fill = header_fill
    for col_idx, portfolio_name in enumerate(weights.columns, start=2):
        cell = ws.cell(row=header_row, column=col_idx, value=portfolio_name)
        cell.font = bold
        cell.fill = header_fill

    row_idx = header_row + 1
    for component, values in weights.iterrows():
        ws.cell(row=row_idx, column=1, value=component)
        for col_idx, value in enumerate(values, start=2):
            cell = ws.cell(row=row_idx, column=col_idx, value=float(value))
            cell.number_format = "0.0%"
        row_idx += 1

    for metric in metric_rows:
        label_cell = ws.cell(row=row_idx, column=1, value=METRIC_LABELS.get(metric, metric))
        label_cell.font = bold
        label_cell.fill = metric_fill
        for col_idx, portfolio_name in enumerate(weights.columns, start=2):
            value = stats.loc[portfolio_name, metric]
            if pd.isna(value):
                excel_value = None
            else:
                excel_value = float(value)
                if metric in PERCENT_POINT_METRICS:
                    excel_value /= 100.0
            cell = ws.cell(row=row_idx, column=col_idx, value=excel_value)
            if metric == "Sharpe":
                cell.number_format = "0.00"
            elif metric in PERCENT_METRICS:
                cell.number_format = "0.0%"
            else:
                cell.number_format = "0.000"
            cell.fill = metric_fill
        row_idx += 1

    ws.freeze_panes = "B4"
    for col_idx in range(1, ws.max_column + 1):
        width = 14
        for values in ws.iter_cols(min_col=col_idx, max_col=col_idx, values_only=True):
            lengths = [len(str(value)) for value in values if value is not None]
            if lengths:
                width = min(36, max(width, max(lengths) + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(output_path)
    return output_path
