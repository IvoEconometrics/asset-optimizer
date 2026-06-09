"""Create machine-friendly Excel input workbooks."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from asset_optimizer.data.loader import ScenarioSet


SPECIAL_COMPONENTS = ("Cashflow", "Benchmark")
GROUP_KEYS = tuple(f"Group_{idx}" for idx in range(1, 6))
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
BOLD = Font(bold=True)


def create_portfolio_input_workbook(
    path: str | Path,
    scenario_set: ScenarioSet,
    *,
    tenor_count: int = 120,
) -> Path:
    """Create the workbook that a portfolio manager fills in."""

    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    _write_settings_sheet(wb, scenario_set)
    _write_cashflow_sheet(wb, scenario_set, tenor_count=tenor_count)
    _write_benchmark_sheet(wb, scenario_set)
    _write_optimization_sheet(wb, scenario_set)
    _write_portfolios_sheet(wb, scenario_set)

    wb.save(output_path)
    return output_path


def _write_settings_sheet(wb: Workbook, scenario_set: ScenarioSet) -> None:
    ws = wb.create_sheet("settings")
    rows = [
        ("Setting", "Value", "Explanation"),
        ("horizon_years", scenario_set.horizon_years, "Minimum horizon is 5 years."),
        ("step_size", 0.025, "Portfolio grid step. Example: 0.025 means 2.5%."),
        ("frontier_points", 250, "Internal frontier sweep count; normally leave unchanged."),
    ]
    _write_rows(ws, rows)
    _style_table(ws, max_col=3)
    _mark_inputs(ws, min_row=2, max_row=4, min_col=2, max_col=2)
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 72


def _write_cashflow_sheet(wb: Workbook, scenario_set: ScenarioSet, *, tenor_count: int) -> None:
    ws = wb.create_sheet("cashflow")
    tenors = scenario_set.yield_tenors or list(range(1, tenor_count + 1))
    _write_rows(ws, [("Tenor", "Weight")])

    for row_idx, tenor in enumerate(tenors, start=2):
        ws.cell(row=row_idx, column=1, value=tenor)
        ws.cell(row=row_idx, column=2, value=0.0)

    _style_table(ws, max_col=2)
    _mark_inputs(ws, min_row=2, max_row=len(tenors) + 1, min_col=2, max_col=2)
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 18


def _write_benchmark_sheet(wb: Workbook, scenario_set: ScenarioSet) -> None:
    ws = wb.create_sheet("benchmark")
    _write_rows(ws, [("Component", "Weight")])

    components = scenario_set.asset_names + ["Cashflow"]
    for row_idx, component in enumerate(components, start=2):
        ws.cell(row=row_idx, column=1, value=component)
        ws.cell(row=row_idx, column=2, value=0.0)

    _style_table(ws, max_col=2)
    _mark_inputs(ws, min_row=2, max_row=len(components) + 1, min_col=2, max_col=2)
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 18


def _write_optimization_sheet(wb: Workbook, scenario_set: ScenarioSet) -> None:
    ws = wb.create_sheet("optimization")
    headers = ("Component", "Min", "Max", *GROUP_KEYS)
    _write_rows(ws, [headers])

    components = scenario_set.asset_names + list(SPECIAL_COMPONENTS)
    for row_idx, component in enumerate(components, start=2):
        ws.cell(row=row_idx, column=1, value=component)

    group_start = len(components) + 4
    _write_rows(ws, [("Group_Key", "Group_Name", "Min", "Max")], start_row=group_start)
    for offset, group_key in enumerate(GROUP_KEYS, start=1):
        row_idx = group_start + offset
        ws.cell(row=row_idx, column=1, value=group_key)

    _style_table(ws, max_col=len(headers))
    _style_header(ws, group_start, max_col=4)
    _mark_inputs(ws, min_row=2, max_row=len(components) + 1, min_col=2, max_col=len(headers))
    _mark_inputs(ws, min_row=group_start + 1, max_row=group_start + len(GROUP_KEYS), min_col=2, max_col=4)
    ws.freeze_panes = "A2"
    _set_widths(ws, [34, 14, 14, 14, 14, 14, 14, 14])


def _write_portfolios_sheet(wb: Workbook, scenario_set: ScenarioSet) -> None:
    ws = wb.create_sheet("portfolios")
    headers = ("Component", "Portfolio_1", "Portfolio_2", "Portfolio_3")
    _write_rows(ws, [headers])

    components = scenario_set.asset_names + list(SPECIAL_COMPONENTS)
    for row_idx, component in enumerate(components, start=2):
        ws.cell(row=row_idx, column=1, value=component)

    _style_table(ws, max_col=len(headers))
    _mark_inputs(ws, min_row=2, max_row=len(components) + 1, min_col=2, max_col=len(headers))
    ws.freeze_panes = "A2"
    _set_widths(ws, [34, 16, 16, 16])


def _write_rows(ws, rows, *, start_row: int = 1) -> None:
    for row_offset, row in enumerate(rows):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=start_row + row_offset, column=col_idx, value=value)


def _style_table(ws, *, max_col: int) -> None:
    _style_header(ws, 1, max_col=max_col)


def _style_header(ws, row: int, *, max_col: int) -> None:
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.font = BOLD
        cell.fill = HEADER_FILL


def _mark_inputs(ws, *, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.fill = INPUT_FILL


def _set_widths(ws, widths: list[int]) -> None:
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
